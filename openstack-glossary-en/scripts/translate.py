"""
OpenStack 용어집 영어 번역 스크립트
한국어 열을 읽어 영어 열의 빈 셀을 채운다. 기존 번역은 유지한다.

사용법:
  python translate.py --file "용어집.xlsx"
  python translate.py --file "용어집.xlsx" --sheets "CON EN,Product"
"""

import argparse
import copy
import sys
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────
# 번역 제외 시트 (무조건 스킵)
# VIOLA: 한국어/영어 구조가 아님 — 절대 처리하지 않는다
# ─────────────────────────────────────────────────────────
EXCLUDED_SHEETS = {"VIOLA"}

# ─────────────────────────────────────────────────────────
# OpenStack 표준 + 공통 UI + 역할 이름 번역 사전
# 새 용어 추가 시 이 딕셔너리에만 추가하면 된다
# ─────────────────────────────────────────────────────────
TRANSLATIONS = {
    # ── LNB / 메뉴 ──────────────────────────────────────
    "대시보드":              "Dashboard",
    "기본 설정":             "Default Settings",
    "이미지":               "Image",
    "키페어":               "Key Pair",
    "인스턴스 유형":          "Flavor",               # OpenStack 공식 용어
    "배치 그룹":             "Server Group",          # OpenStack Affinity/Anti-Affinity
    "볼륨 그룹 타입":         "Volume Group Type",
    "컴퓨트":               "Compute",
    "인스턴스":              "Instance",
    "인스턴스 스냅샷":        "Instance Snapshot",
    "VPC":                "VPC",
    "세그먼트":              "Segment",
    "라우터":               "Router",
    "유동 IP":              "Floating IP",
    "로드밸런서":             "Load Balancer",
    "보안 그룹":             "Security Group",
    "스토리지":              "Storage",
    "볼륨":                 "Volume",
    "볼륨 스냅샷":           "Volume Snapshot",
    "볼륨 타입":             "Volume Type",
    "볼륨 그룹":             "Volume Group",
    "볼륨 그룹 스냅샷":       "Volume Group Snapshot",
    "볼륨 백업":             "Volume Backup",
    "볼륨 백업 스케줄":       "Volume Backup Schedule",
    "공유 파일":             "Shared File",           # OpenStack Manila 기반
    "오브젝트":              "Object Storage",        # OpenStack Swift 기반
    "모니터링":              "Monitoring",
    "랙 구성도":             "Rack Diagram",
    "물리 네트워크 구성도":    "Physical Network Topology",
    "가상 네트워크 구성도":    "Virtual Network Topology",
    "HA 관리":              "HA Management",
    "HA 설정":              "HA Settings",
    "HA 호스트":             "HA Host",
    "관리":                 "Management",
    "프로젝트":              "Project",
    "사용자":               "User",
    "호스트":               "Host",
    "호스트 그룹":           "Host Group",
    "스토리지 백엔드":        "Storage Backend",
    "SSL 인증서":            "SSL Certificate",
    "임계치 알람":           "Threshold Alarm",
    "알람 채널 관리":         "Alarm Channel Management",
    "시스템 정보":            "System Information",
    "시스템 점검":            "System Maintenance",
    "IP 접근 제어":          "IP Access Control",
    "Product 엔진":      "Product Engine",

    # ── 공통 UI ─────────────────────────────────────────
    "작업":                 "Actions",
    "생성일시":              "Created At",
    "생성자":               "Created By",
    "사용량":               "Usage",
    "사용률":               "Usage Rate",
    "전체 수 {n}":          "Total {n}",
    "엑셀 다운로드":          "Download Excel",
    "이름을 입력해 주세요.":   "Please enter a name.",
    "중복확인":              "Duplicate Check",

    # ── CON 전용 ────────────────────────────────────────
    "마이그레이션":           "Migration",
    "다중 마이그레이션":       "Bulk Migration",
    "허용 IP":              "Allowed IP",
    "차단 IP":              "Blocked IP",
    "버킷/컨테이너":          "Bucket/Container",
    "백업":                 "Backup",
    "가용성 존":             "Availability Zone",
    "테넌트":               "Tenant",
    "임계치":               "Threshold",
    "로그":                 "Log",

    # ── 역할 이름 ────────────────────────────────────────
    "최고관리자":             "Super Admin",
    "계정관리자":             "Account Admin",
    "감사자":               "Auditor",
    "솔루션 관리자":          "Solution Admin",
    "Product 엔진 관리자": "Product Engine Admin",
    "기본설정 관리자":         "Basic Settings Admin",
    "물리 인프라 관리자":      "Physical Infra Admin",
    "컴퓨트 관리자":          "Compute Admin",
    "네트워크 관리자":         "Network Admin",
    "스토리지 관리자":         "Storage Admin",
    "프로젝트 관리자":         "Project Admin",
}


def find_kor_eng_columns(ws):
    """
    시트에서 한국어 열(kor_col)과 영어 열(eng_col)의 인덱스(1-based)를 반환한다.
    헤더 행에서 '한국어' 셀을 찾고, 바로 오른쪽을 영어 열로 간주한다.
    찾지 못하면 (None, None) 반환.
    """
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value == "한국어":
                return cell.column, cell.column + 1
    return None, None


def fill_sheet(ws, kor_col, eng_col):
    """
    한국어 열을 읽어 영어 열의 빈 셀만 채운다.
    Returns: (채운 수, 사전 미등록 항목 리스트)
    """
    filled = 0
    missing = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        kor_cell = row[kor_col - 1]
        eng_cell = row[eng_col - 1]

        kor_val = kor_cell.value
        if not kor_val or eng_cell.value:
            continue  # 빈 한국어 셀이거나 이미 영어가 있으면 건너뜀

        eng_val = TRANSLATIONS.get(str(kor_val).strip())
        if eng_val:
            if kor_cell.font:
                eng_cell.font = copy.copy(kor_cell.font)
            if kor_cell.alignment:
                eng_cell.alignment = copy.copy(kor_cell.alignment)
            eng_cell.value = eng_val
            filled += 1
        else:
            missing.append(str(kor_val).strip())

    return filled, missing


def translate_file(filepath, target_sheets=None):
    """
    파일 전체(또는 지정 시트)에 번역을 적용하고 저장한다.
    Returns: (총 채운 수, {시트명: [미등록 항목 리스트]})
    """
    wb = load_workbook(filepath)
    sheets = target_sheets if target_sheets else wb.sheetnames

    total_filled = 0
    all_missing = {}

    for sheet_name in sheets:
        if sheet_name in EXCLUDED_SHEETS:
            print(f"[{sheet_name}] 제외 시트 — 건너뜀")
            continue

        if sheet_name not in wb.sheetnames:
            print(f"[경고] 시트 없음: '{sheet_name}'")
            continue

        ws = wb[sheet_name]
        kor_col, eng_col = find_kor_eng_columns(ws)

        if kor_col is None:
            print(f"[{sheet_name}] '한국어' 헤더를 찾을 수 없어 건너뜀")
            continue

        filled, missing = fill_sheet(ws, kor_col, eng_col)
        total_filled += filled
        if missing:
            all_missing[sheet_name] = missing

        print(f"[{sheet_name}] {filled}개 입력 완료"
              + (f", 미등록 {len(missing)}개" if missing else ""))

    wb.save(filepath)
    print(f"\n총 {total_filled}개 번역 입력 / 파일 저장: {filepath}")

    if all_missing:
        print("\n[사전 미등록 항목 — 수동 확인 필요]")
        for sheet, items in all_missing.items():
            for item in items:
                print(f"  [{sheet}] {item}")

    return total_filled, all_missing


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="OpenStack 용어집 영어 번역 자동 입력"
    )
    parser.add_argument("--file", required=True, help="Excel 파일 경로")
    parser.add_argument(
        "--sheets",
        default="",
        help="대상 시트명 (쉼표 구분). 생략 시 전체 시트 적용",
    )
    args = parser.parse_args()

    sheets = [s.strip() for s in args.sheets.split(",") if s.strip()] or None
    total, missing = translate_file(args.file, sheets)
    sys.exit(0 if not missing else 1)
